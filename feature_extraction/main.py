# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import os
from soundtype_params import SoundtypeParams
import numpy as np
import xlrd
import pandas as pd

# import xlwt

'''
1 M208 0746(0-10s)
2 M609 1147(0-10s)
3 M112 0800(5-15s)
4 M112 0800(50-60s)
5 M209 0820(40-50s)
6 M209 0820(60s-70s)
7  M202 1415(130-140s)
8  M202 1415(270-280s)
9  M204 1446(240-250s)
10 N401 0900(0-10s)
11  N601 0847(265-275s)
12  N402 1744(125-135s)
'''

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # wb = xlrd.open_workbook('Record106377.xls')
    # 1. 修改路径
    path = r'E:\Python\Code\my_noise_recognize\feature_extraction\input\N118-1630.xls'
    start_time = 0
    end_time = 10
    wb = xlrd.open_workbook(path)
    params_dict = {}
    for idx in range(6):
        sheet = wb.sheet_by_index(idx)
        # 读取同一个excel文件中的不同sheet
        print("Sheet", idx + 1)
        param_type = sheet.cell(10, 1).value
        if param_type == 'pressure':
            param_type = 'Leq'
        elif param_type == 'fluctuation strength':
            param_type = 'Fluct'
        else:
            # 不用小写
            # param_type = param_type.lower()
            # 首字母大写
            param_type = param_type.capitalize()
        values = []
        times = []

        # 将文件中所有数据都读取出来
        for i in range(15, sheet.nrows):
            time = float(sheet.cell(i, 0).value)
            # 2. 筛选时间片段
            if float(start_time) <= time < float(end_time):
                values.append(float(sheet.cell(i, 1).value))
                times.append(time)
        '''
        M208 0746(0-10s)
        M609 1147(0-10s)
        M112 0800(5-15s)
        M112 0800(50-60s)
        M209 0820(40-50s)
        M209 0820(60s-70s)
        M202 1415(130-140s)
        M202 1415(270-280s)
        M204 1446(240-250s)
        N401 0900(0-10s)
        N601 0847(265-275s)
        N402 1744(125-135s)
        '''
        values = np.array(values)
        times = np.array(times)
        mean = values.mean()
        var = values.var()
        std = values.std(ddof=1)
        max = values.max()
        p_10 = np.percentile(values, 10, interpolation='midpoint')
        p_25 = np.percentile(values, 25, interpolation='midpoint')
        p_50 = np.percentile(values, 50, interpolation='midpoint')
        p_75 = np.percentile(values, 75, interpolation='midpoint')
        p_90 = np.percentile(values, 90, interpolation='midpoint')
        p_10_90 = p_90 - p_10

        params_dict["%s_mean" % param_type] = mean
        params_dict["%s_var" % param_type] = var
        params_dict["%s_std" % param_type] = std
        params_dict["%s_max" % param_type] = max
        params_dict["%s_10" % param_type] = p_10
        params_dict["%s_25" % param_type] = p_25
        params_dict["%s_median" % param_type] = p_50
        params_dict["%s_75" % param_type] = p_75
        params_dict["%s_90" % param_type] = p_90
        params_dict["%s_10_90" % param_type] = p_10_90

    soundtype_params = None
    try:
        file_name = os.path.basename(path).replace("xlsx", "wav")
        print(file_name, "is processed !")
        leq_mean = params_dict["Leq_mean"]
        leq_std = params_dict["Leq_std"]
        leq_25 = params_dict["Leq_25"]
        leq_median = params_dict["Leq_median"]
        leq_75 = params_dict["Leq_75"]
        leq_10_90 = params_dict["Leq_10_90"]
        loudness_mean = params_dict["Loudness_mean"]
        loudness_std = params_dict["Loudness_std"]
        loudness_25 = params_dict["Loudness_25"]
        loudness_median = params_dict["Loudness_median"]
        loudness_75 = params_dict["Loudness_75"]
        loudness_10_90 = params_dict["Loudness_10_90"]
        roughness_mean = params_dict["Roughness_mean"]
        roughness_std = params_dict["Roughness_std"]
        roughness_25 = params_dict["Roughness_25"]
        roughness_median = params_dict["Roughness_median"]
        roughness_75 = params_dict["Roughness_75"]
        roughness_10_90 = params_dict["Roughness_10_90"]
        sharpness_mean = params_dict["Sharpness_mean"]
        sharpness_std = params_dict["Sharpness_std"]
        sharpness_25 = params_dict["Sharpness_25"]
        sharpness_median = params_dict["Sharpness_median"]
        sharpness_75 = params_dict["Sharpness_75"]
        sharpness_10_90 = params_dict["Sharpness_10_90"]
        fluct_mean = params_dict["Fluct_mean"]
        fluct_std = params_dict["Fluct_std"]
        fluct_25 = params_dict["Fluct_25"]
        fluct_median = params_dict["Fluct_median"]
        fluct_75 = params_dict["Fluct_75"]
        fluct_10_90 = params_dict["Fluct_10_90"]
        tonality_mean = params_dict["Tonality_mean"]
        tonality_std = params_dict["Tonality_std"]
        tonality_25 = params_dict["Tonality_25"]
        tonality_median = params_dict["Tonality_median"]
        tonality_75 = params_dict["Tonality_75"]
        tonality_10_90 = params_dict["Tonality_10_90"]

        soundtype_params = SoundtypeParams(file_name, leq_mean, leq_std, leq_25, leq_median, leq_75, leq_10_90,
                                           loudness_mean, loudness_std, loudness_25, loudness_median, loudness_75,
                                           loudness_10_90, roughness_mean, roughness_std, roughness_25,
                                           roughness_median, roughness_75, roughness_10_90, sharpness_mean,
                                           sharpness_std, sharpness_25, sharpness_median, sharpness_75, sharpness_10_90,
                                           fluct_mean, fluct_std, fluct_25, fluct_median, fluct_75, fluct_10_90,
                                           tonality_mean, tonality_std, tonality_25, tonality_median, tonality_75,
                                           tonality_10_90)
        # print("aaa", soundtype_params)
    except Exception as e:
        print("[ERROR] XlsProcessor: %s" % e)

    params_dict["name"] = os.path.basename(path)
    a = pd.DataFrame(params_dict, [0])

    write_path = r"E:\Python\Code\my_noise_recognize\feature_extraction\output\feature_extraction.xlsx"

    from openpyxl import load_workbook

    # # excel追加写入（xlsx：批注不会丢失）
    # def write_excel_xlsx_append(file_path, df):  # df:DataFrame
    #     wb = load_workbook(file_path)  # type:workbook
    #     ws = wb.active  # type:worksheet
    #     for i in range(0, len(df)):  # 行
    #         for j, c in enumerate("ABCDEFGHIJKLM"):  # 列
    #             ws[f"{c}{i + 3}"] = str(df.iloc[i][j])  # 追加数据，从3行开始写入
    #     wb.save(file_path)

    writer = pd.ExcelWriter(write_path, mode='a', engine="openpyxl")
    # writer = pd.ExcelWriter(write_path)
    a.to_excel(writer, float_format='%.5f')
    writer.save()
    writer.close()

    # See PyCharm help at https://www.jetbrains.com/help/pycharm/
